#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Dec  5 23:20:49 2017

@author: aldg
"""

#constantes que no se pueden iniciar en el notebook 

T_ENUNCIADO=2
T_ENUNCIADO_SIGNO=3
T_ENUNCIADO_BORDE_INFERIOR=5
T_ENUNCIADO_BORDE_INFERIOR_IZQ=6
T_FC_RESULTADO=10
T_FC_RESULTADO_RANGE=20

T_ENUNCIADO_RESTA_DIVI=8  #el - de las restas auxiliares de las divisiones
T_FC_RESULTADO_RESTA_DIVI=12  #el borde inforrior de las restas auxiliares de las divisiones
T_FC_RESULTADO_SIGNO_RESTA_DIVI=13  #desinvisibiliza el signo menos de las restas auxiliares de las divisiones

T_FC_ESCRITO_FUERA=30  # escribe fuera de las casillas de la operacion


from openpyxl.styles import Font, PatternFill, Border, Side,Alignment

color_fondo_defecto='FFFFFF'   #blanco
color_tinta_defecto='000000'   #negro

formatos=None

def func_formatos ():
    fmt = {
    #


        'tinta_invisible':  { 'bold': True,
                            'size': CFG_FONT_SIZE,
                            'color':color_fondo_defecto },     # T_ENUNCIADO_RESTA_DIVI  >-invisibe
        'tinta_defecto':  { 'bold': True,
                            'size': CFG_FONT_SIZE, 
                            'color':color_tinta_defecto },     # vivibele T_FC_RESULTADO_SIGNO_RESTA_DIVI #desinvisibiliza el signo menos de las restas auxiliares de las divisiones

        'respuesta_intermedia_ok': {'font':  Font(bold=True, size=CFG_FONT_SIZE, color=CFG_COLOR_INTERMEDIO_OK)},  # respuesta intermedia OK # type: ignore
        'respuesta_final_ok': {
            'font':  Font(bold=True, size=CFG_FONT_SIZE, color=color_tinta_defecto),
            'fill':  PatternFill("solid", bgColor=CFG_COLOR_FINAL_OK)},  # respuesta intermedia OK
        'respuesta_intermedia_ko': {'fill':  PatternFill("solid", bgColor=CFG_COLOR_INTERMEDIO_KO )},  # respuesta final KO
        'cuadritos_linea_superior': {
            'font':  Font(bold=True, size=CFG_FONT_SIZE, color=CFG_COLOR_FINAL_OK),
            'fill':  PatternFill("solid", bgColor=CFG_COLOR_FINAL_OK)},  # respuesta intermedia OK
        'respuesta_escrita_fuera': {'fill':  PatternFill("solid", bgColor=CFG_COLOR_ESCRITO_FUERA )},  # escrito fuera de sitio rojo claro FF6666

        'enunciado_borde_inferior': {
                "bottom": Side(border_style='thick', color=color_tinta_defecto)}, #T_ENUNCIADO_BORDE_INFERIOR
        'enunciado_borde_inferior_izq': {
                "bottom": Side(border_style='thick', color=color_tinta_defecto),
                "left":   Side(border_style='thick', color=color_tinta_defecto) },      #T_ENUNCIADO_BORDE_INFERIOR_IZQ      
        'enunciado_borde_inferior_resta_div': {
                "bottom":Side(border_style='medium', color=color_tinta_defecto)}, #T_ENUNCIADO_RESTA_DIVI (linea auxiliar de las restas haciendo divisiones)

        'backupxxxx_respuesta_intermedia_ok': {'fill': PatternFill("solid", bgColor =CFG_COLOR_INTERMEDIO_OK)},  # respuesta intermedia OK
        'key2': 'value2',
        'key3': 'value3'
    }
    return(fmt)
#        fill_color_interOk=PatternFill("solid", bgColor =CFG_COLOR_INTERMEDIO_OK)
