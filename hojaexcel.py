#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Dec  5 00:10:10 2017

@author: aldg
"""

from constantes import *

from operacion import Regla, Operacion,Suma,Resta,Multiplicacion,Division,Division_con_restas

from openpyxl import Workbook
#from openpyxl.compat import range
from openpyxl.utils import get_column_letter

from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border, Side,Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
# from openpyxl.styles.colors import RED,GREEN,DARKGREEN,DARKRED
from openpyxl.worksheet.datavalidation import DataValidation
from math import ceil


# import builtins


# llamoa a calcula_size con el tamaño de la 1a version  (11 es el tamaño de la 1a version de fontsize)
def calc_size(sz):
    return (round(sz*CFG_FONT_SIZE/11))

#evita que se seleccione por defecto un modo no disponible,si no existe ,elige el primero disponible
def modo_defecto():
    if CFG_MODO_SELEC in CFG_MODOS_DISPO.split(',') :
        return CFG_MODO_SELEC
    else:
        return CFG_MODOS_DISPO.split(',')[0]
    


class HojaExcel:
    def __init__ (self, nombre, operaciones_por_fila=3 ):
        self.nombre = nombre
        self.operaciones_por_fila = operaciones_por_fila
        self.celdas_por_operacion = None
        self.numero_operaciones = None


        self.operaciones=[]
        global formatos
        formatos=func_formatos()   #para evitar problema constantes dinamicas en segundas reejecuciones de collab jupyter

    def inicializa(self):



        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Operaciones"

        self.celdanivel="$N$2"
        self.celdanivelcombo="$H$2"
        



        self.ws1.freeze_panes='A3'

        self.ws1.row_dimensions[2].height=calc_size(24)  
        max_x=max(46,self.operaciones_por_fila*self.celdas_por_operacion+CFG_MARGEN_X*2)  #max 46 para que cuadren los titulos
        max_y=self.celdas_por_operacion*(ceil(self.numero_operaciones/self.operaciones_por_fila))+CFG_MARGEN_Y*2+10

        for n in range(3,max_y):
            self.ws1.row_dimensions[n].height=calc_size(12+3) #5 margen extra para prblema al editar con  excel que no deja ver la celda de encima
 
        for n in range(1,max_x):
            self.ws1.column_dimensions[ get_column_letter(n)].width=(2+6/8)*CFG_FONT_SIZE/11 # 3/8 margen exrra para problema excel unidades width =caracteres*8+5
            #self.ws1.column_dimensions[get_column_letter(n)].font =Font(bold=True)


        for y in range(1,max_y):
            for x in range(1, max_x):
                self.ws1.cell(row=y, column=x).alignment  = Alignment(horizontal="center", vertical="center")
                self.ws1.cell(row=y, column=x).font=Font(bold=True,size=CFG_FONT_SIZE)
# al = Alignment(horizontal="center", vertical="center")


#         style = Style(font )
# for col in 'ABCD':
#      ws._styles[col] = style





        celdanivelcombo=self.celdanivelcombo
        celdanivel = self.celdanivel

        self.ws1.merge_cells('C2:G2')
        self.ws1["C2"].font=Font(bold=True,size=calc_size(14),color="0066FF")
        self.ws1["C2"]="MODO = "

        self.ws1.merge_cells('H2:M2')

        dv = DataValidation(type="list", formula1='"'+CFG_MODOS_DISPO+'"', allow_blank=False,showDropDown=False)
        self.ws1.add_data_validation(dv)

        self.ws1[celdanivelcombo].font=Font(bold=True,size=calc_size(15),color="0066FF")
        self.ws1[celdanivelcombo] =modo_defecto()
        dv.add(self.ws1[celdanivelcombo])

        self.ws1[celdanivel]='=IF('+celdanivelcombo+'="Fácil",0,IF('+celdanivelcombo+'= "Pro",1,IF('+celdanivelcombo+'= "SuperPro",2,"")))'
        self.ws1[celdanivel].font =Font(**formatos['tinta_invisible'])

# RESUELTAS
        self.ws1.merge_cells('P2:V2')
        self.ws1["P2"].font=Font(bold=True,size=calc_size(14),color="04B431")
        self.ws1["P2"]="RESUELTAS ="

        self.ws1.merge_cells('W2:Z2')
        self.ws1["W2"].font = Font(bold=True, size=calc_size(20), color="04B431")
        self.ws1["W2"]='=COUNTIF(1:1,"=1")'

#faltan
        self.ws1.merge_cells('AA2:AF2')
        self.ws1["AA2"].font=Font(bold=True,size=calc_size(14),color="DF0101")
        self.ws1["AA2"]="FALTAN ="

        self.ws1.merge_cells('AG2:AJ2')
        self.ws1["AG2"].font = Font(bold=True, size=calc_size(20), color="DF0101")
        self.ws1["AG2"]='=COUNTIF(1:1,"=0")'

# =CONTAR.SI(1:1;"=0")



    def add_operacion(self,ope):
        self.operaciones.append(ope)




    def carga_operaciones(self):

        self.numero_operaciones=len(self.operaciones)

        # calcula CELDAS_POR_OPERACION (12 defecto acterior)

        coordenadas_regl_ope= [  [ (regl.posx,regl.posy,regl.posx+len(str( regl.valor))-1 ) for  regl in ope.reglas if regl.tipo!=T_FC_ESCRITO_FUERA ] for ope in self.operaciones ]
        # (x para minimos,y, x para maximos) por ope

        width_max = max  (   max( regl[2] for regl in ope) - min( regl[0]  for regl in ope) +1  for ope in coordenadas_regl_ope)
        height_max= max  (   max( regl[1] for regl in ope) - min( regl[1]  for regl in ope) +1  for ope in coordenadas_regl_ope)
        # debug=list(width_ope)
        # height_debug=list(height_ope)
            
        self.celdas_por_operacion=max(width_max+2,height_max+2,CFG_MIN_CELDAS_POR_OPERACION)

        # fin calcula CELDAS_POR_OPERACION (12 defecto acterior)
        self.inicializa()

        self.x_operacion=0
        self.y_operacion=0

        for ope in self.operaciones:
            self.carga_operacion(ope)
            self.x_operacion=(self.x_operacion+1)% self.operaciones_por_fila
            if self.x_operacion==0 :
                self.y_operacion+=1
            print (ope.__class__ ,ope.terminos,"--> OK")

    def carga_operacion(self,ope):

        #greenFill = PatternFill("solid", bgColor ="E0F8E6")
        #colordarkgreen="2EFE2E"
        #darkgreenFill = PatternFill("solid", bgColor = colordarkgreen)


        # color_finalOK=CFG_COLOR_FINAL_OK
        # fill_color_finalOK=PatternFill("solid", bgColor = color_finalOK)
#        fill_color_interOk=PatternFill("solid", bgColor =CFG_COLOR_INTERMEDIO_OK)
         



        desp_x=self.x_operacion*self.celdas_por_operacion+CFG_MARGEN_X
        desp_y=self.y_operacion*self.celdas_por_operacion+CFG_MARGEN_Y
        izqui=ope.GetIzquierda()

        tx,ty=ope.GetTamano()
        desp_x=desp_x+int((self.celdas_por_operacion-tx)/2)-izqui
        desp_y=desp_y+int((self.celdas_por_operacion-ty)/2)




        # T_ENUNCIADO,T_ENUNCIADO_SIGNO
        for n in [ r for r in ope.reglas if r.tipo in(T_ENUNCIADO,T_ENUNCIADO_SIGNO)  ]:
            self.ws1.cell(column=n.posx+desp_x, row=n.posy+desp_y,value=n.valor)

        # T_ENUNCIADO_BORDE_INFERIOR

        for n in [ r for r in ope.reglas if r.tipo==T_ENUNCIADO_BORDE_INFERIOR]:
            self.ws1.cell(column=n.posx + desp_x, row=n.posy + desp_y).border=Border( **formatos['enunciado_borde_inferior'])

        # T_ENUNCIADO_BORDE_INFERIOR_IZQ

        for n in [ r for r in ope.reglas if r.tipo==T_ENUNCIADO_BORDE_INFERIOR_IZQ]:
            self.ws1.cell(column=n.posx + desp_x, row=n.posy + desp_y).border=Border( **formatos['enunciado_borde_inferior_izq'])


        # T_FC_RESULTADO_RANGE
        # darkgreenFill = PatternFill(start_color=RED , end_color=RED,
        #                         fill_type='solid')



        formula = "AND("
        rango = []
        for n in [r for r in ope.reglas if r.tipo == T_FC_RESULTADO_RANGE]:

            if formula != "AND(":
                formula+=","

            rango.append("$" + get_column_letter(n.posx + desp_x) + "$" + str(n.posy + desp_y) \
                   + ":$" + get_column_letter(n.posx + desp_x + len(n.valor) - 1) + "$" + str(n.posy + desp_y))
            # print( rango)

            for c in range(n.posx, n.posx + len(n.valor) ):
                if formula!="AND(" and formula[-1]!=',' :
                    formula+="&"
                formula += "$" + get_column_letter(c +desp_x) + "$" + str(n.posy +desp_y)
            formula += '&"" ="' + str(n.valor) + '"'   # le concateno un texto vacio "" para convertirlo a texto cuando solo hay un digito
            # print(formula)
        formula +=")"   # cierro el AND de la formula excel
        # print(formula)

        celdaformula="$"+get_column_letter(self.y_operacion*self.operaciones_por_fila+ self.x_operacion +1)+"$1"
        self.ws1[celdaformula]="=IF("+formula+",1,0)"
        self.ws1[celdaformula].font=Font(**formatos['tinta_invisible'])
        self.ws1.conditional_formatting.add(celdaformula,   FormulaRule(formula=[celdaformula+"=1"],**formatos['cuadritos_linea_superior']))


        formula="$"+get_column_letter(self.y_operacion*self.operaciones_por_fila+ self.x_operacion +1)+"$1=1"
        formula = 'AND(' + self.celdanivel + '<2,' + formula + ')'
        for rg in rango:
          self.ws1.conditional_formatting.add(rg,
                                           FormulaRule(formula=[formula],stopIfTrue=True, **formatos['respuesta_final_ok']))

        # T_FC_RESULTADO

        for n in [ r for r in ope.reglas if r.tipo==T_FC_RESULTADO ]:
             formula="$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y)+'= '+str(n.valor)
             if n.valor==0:
                 formula='AND(NOT(ISBLANK('+"$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y)+')),'+formula+')'

             formula='AND('+self.celdanivel +'=0,'+formula+')'

             self.ws1.conditional_formatting.add("$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y),FormulaRule(formula=[formula], **formatos['respuesta_intermedia_ok']))

             # print("formula: "+"$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y),formula)
             #Fomato para resltado intermedio mal
             formula="$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y)+'<> '+str(n.valor)
             formula='AND(NOT(ISBLANK('+"$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y)+')),'+formula+')'
             formula='AND('+self.celdanivel +'=0,'+formula+')'
             self.ws1.conditional_formatting.add("$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y),FormulaRule(formula=[formula], **formatos['respuesta_intermedia_ko']))


        # T_ENUNCIADO_RESTA_DIVI  >-invisibe , lo visibiliza la condicion T_FC_RESULTADO_SIGNO_RESTA_DIVI
        for n in [ r for r in ope.reglas if r.tipo ==(T_ENUNCIADO_RESTA_DIVI)  ]:
            self.ws1.cell(column=n.posx+desp_x, row=n.posy+desp_y,value=n.valor)
            self.ws1.cell(column=n.posx+desp_x, row=n.posy+desp_y).font=Font(**formatos['tinta_invisible'])






        # T_FC_RESULTADO_SIGNO_RESTA_DIVI #desinvisibiliza el signo menos de las restas auxiliares de las divisiones
        # si CFG_LINEAS_RESTA_AUX_DIVI_FIJAS=True => siempre visibles
        for n in [ r for r in ope.reglas if r.tipo ==(T_FC_RESULTADO_SIGNO_RESTA_DIVI)  ]:
             if CFG_LINEAS_RESTA_AUX_DIVI_FIJAS:
                 formula="TRUE"
             else:    
                formula="$"+get_column_letter(n.test_posx+desp_x)+"$"+str(n.test_posy+desp_y)+'= '+str(n.test_valor)
                if n.valor==0:
                    formula='AND(NOT(ISBLANK('+"$"+get_column_letter(n.test_posx+desp_x)+"$"+str(n.test_posy+desp_y)+')),'+formula+')'
                # formula='AND('+self.celdanivel +'=0,'+formula+')'

             self.ws1.conditional_formatting.add("$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y),FormulaRule(formula=[formula], font=Font(**formatos['tinta_defecto'])))
            
        # T_FC_RESULTADO_RESTA_DIVI  #visibiliza la linea de resta cuando va resolviendose el cociente
        # si CFG_LINEAS_RESTA_AUX_DIVI_FIJAS=True => siempre visibles
        for n in [ r for r in ope.reglas if r.tipo ==(T_FC_RESULTADO_RESTA_DIVI)  ]:
             if CFG_LINEAS_RESTA_AUX_DIVI_FIJAS:
                 formula="TRUE"
             else:    
                formula="$"+get_column_letter(n.test_posx+desp_x)+"$"+str(n.test_posy+desp_y)+'= '+str(n.test_valor)
                if n.valor==0:
                    formula='AND(NOT(ISBLANK('+"$"+get_column_letter(n.test_posx+desp_x)+"$"+str(n.test_posy+desp_y)+')),'+formula+')'
                # formula='AND('+self.celdanivel +'=0,'+formula+')'


             rg="$" + get_column_letter(n.posx + desp_x) + "$" + str(n.posy + desp_y) \
                   + ":$" + get_column_letter(n.posx + desp_x + len(n.valor) - 1) + "$" + str(n.posy + desp_y)
             self.ws1.conditional_formatting.add(rg,FormulaRule(
                 formula=[formula], border=Border( **formatos['enunciado_borde_inferior_resta_div'])  )  )
             

        # T_FC_ESCRITO_FUERA  >-escrito en casillas fuera de la operacion , No se debe escribir nada ,por ejemplo debajo del cociente
        for n in [ r for r in ope.reglas if r.tipo ==(T_FC_ESCRITO_FUERA)  ]:
             formula="$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y)+'<> '+str(n.valor)
             formula='AND(NOT(ISBLANK('+"$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y)+')),'+formula+')'
             formula='AND('+self.celdanivel +'=0,'+formula+')'    
             self.ws1.conditional_formatting.add("$"+get_column_letter(n.posx+desp_x)+"$"+str(n.posy+desp_y),FormulaRule(formula=[formula], **formatos['respuesta_escrita_fuera']))


    def graba_hoja(self):
        self.wb.save(filename = self.nombre)

